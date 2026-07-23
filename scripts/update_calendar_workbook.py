from __future__ import annotations

from pathlib import Path
from datetime import datetime
import json
import os
import shutil
import subprocess
import sys

from openpyxl import load_workbook

PROJECT = Path(__file__).resolve().parent
ROOT = Path(os.environ.get('DASHBOARD_REPO_DIR', PROJECT.parents[0]))
SOURCE_XLSX = Path(os.environ.get('SOURCE_XLSX', ROOT / 'data' / 'holiday_student_analysis.xlsx'))
CANDIDATES = PROJECT / 'official_calendar_cache' / 'candidates.json'
BUILDER = PROJECT / 'create_holiday_dashboard.py'
GITHUB_PAGES_URL = 'https://yxq870504-lgtm.github.io/SXZX-FJKXSJ/'
GRADE_ORDER = ['初一', '初二', '初三', '高一', '高二', '高三']


def norm_province(v):
    s = str(v or '').strip()
    for suf in ['维吾尔自治区', '壮族自治区', '回族自治区', '特别行政区', '自治区', '省', '市']:
        s = s.replace(suf, '')
    return s


def run_git(args, check=True):
    proc = subprocess.run(
        ['git', '-C', str(ROOT), *args],
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        check=False,
    )
    if check and proc.returncode != 0:
        detail = (proc.stderr or proc.stdout or '').strip()
        raise RuntimeError(f"git {' '.join(args)} 失败：{detail}")
    return proc


def publish_to_github(summary: str):
    if not (ROOT / '.git').exists():
        return {'git': 'skipped', 'message': f'{ROOT} 不是 Git 仓库'}
    run_git(['config', 'user.name', 'github-actions[bot]'], check=False)
    run_git(['config', 'user.email', '41898282+github-actions[bot]@users.noreply.github.com'], check=False)
    run_git(['add', 'index.html', 'data/holiday_student_analysis.xlsx'])
    status = run_git(['status', '--porcelain'])
    if not status.stdout.strip():
        return {'git': 'unchanged', 'message': 'index.html 和底表无变化，无需推送'}
    msg = f'daily official calendar refresh {datetime.now():%Y-%m-%d %H:%M:%S} {summary}'
    run_git(['commit', '-m', msg])
    run_git(['push'])
    return {'git': 'pushed', 'message': msg, 'url': GITHUB_PAGES_URL}


def ensure_sheets(wb):
    if '官方校历刷新日志' not in wb.sheetnames:
        ws = wb.create_sheet('官方校历刷新日志')
        ws.append(['刷新时间', '候选数', '自动更新数', '保留预估数', 'Git同步状态', '备注'])
    if '官方校历命中明细' not in wb.sheetnames:
        ws = wb.create_sheet('官方校历命中明细')
        ws.append(['刷新时间', '省份', '放假日期', '开学日期', '置信度', '来源URL', '处理结果'])


def main():
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f'未找到 Excel 底表：{SOURCE_XLSX}')
    if not CANDIDATES.exists():
        raise FileNotFoundError(f'未找到候选文件：{CANDIDATES}')

    data = json.loads(CANDIDATES.read_text(encoding='utf-8'))
    candidates = data.get('candidates', [])
    best = {}
    for c in candidates:
        if float(c.get('confidence') or 0) < 0.85:
            continue
        p = norm_province(c.get('province'))
        prev = best.get(p)
        if not prev or float(c.get('confidence') or 0) > float(prev.get('confidence') or 0):
            best[p] = c

    wb = load_workbook(SOURCE_XLSX)
    ensure_sheets(wb)
    ws = wb['省份日历_人数底表']
    province_to_row = {}
    for r in range(2, ws.max_row + 1):
        province_to_row[norm_province(ws.cell(r, 2).value)] = r

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updated = 0
    for pkey, c in best.items():
        if pkey not in province_to_row:
            continue
        h = c.get('holiday_date')
        sp = c.get('spring_date')
        if not h and not sp:
            continue
        row = province_to_row[pkey]
        changed = False
        for gi, g in enumerate(GRADE_ORDER):
            base_col = 3 + gi * 5
            if h:
                ws.cell(row, base_col).value = h
                changed = True
            if sp:
                ws.cell(row, base_col + 1).value = sp
                changed = True
            if changed:
                ws.cell(row, base_col + 2).value = '是'
        if changed:
            updated += 1
            wb['官方校历命中明细'].append([now, c.get('province'), h or '', sp or '', c.get('confidence'), c.get('source_url'), '已自动更新'])

    keep_estimated = len(province_to_row) - updated
    wb['官方校历刷新日志'].append([now, len(candidates), updated, keep_estimated, 'pending', '准备生成看板并同步 GitHub'])

    backup_dir = PROJECT / 'backups'
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f'{SOURCE_XLSX.stem}_official_refresh_{datetime.now():%Y%m%d_%H%M%S}.xlsx'
    shutil.copy2(SOURCE_XLSX, backup)
    wb.save(SOURCE_XLSX)

    subprocess.run([sys.executable, str(BUILDER)], cwd=str(PROJECT), check=True)
    publish = publish_to_github(f'updated={updated} candidates={len(candidates)}')
    print(json.dumps({'ok': True, 'candidates': len(candidates), 'updated': updated, 'keep_estimated': keep_estimated, 'publish': publish, 'backup': str(backup)}, ensure_ascii=False))


if __name__ == '__main__':
    main()
