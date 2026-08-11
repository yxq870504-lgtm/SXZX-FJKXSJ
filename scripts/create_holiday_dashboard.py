from pathlib import Path
import os
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import json

repo_root = Path(os.environ.get('DASHBOARD_REPO_DIR', Path(__file__).resolve().parents[1]))
source_file = Path(os.environ.get('SOURCE_XLSX', repo_root / 'data' / 'holiday_student_analysis.xlsx'))
repo_dir = repo_root
repo_dir.mkdir(parents=True, exist_ok=True)
out_file = repo_dir / 'index.html'

GRADE_ORDER = ['初一', '初二', '初三', '高一', '高二', '高三']
STAGE_GRADES = {'初中': ['初一', '初二', '初三'], '高中': ['高一', '高二', '高三']}

SPECIAL_SHORT = {
    '内蒙古自治区': '内蒙古', '广西壮族自治区': '广西', '宁夏回族自治区': '宁夏',
    '新疆维吾尔自治区': '新疆', '西藏自治区': '西藏', '黑龙江省': '黑龙江',
    '香港特别行政区': '香港', '澳门特别行政区': '澳门'
}

def short_province(name):
    if not name:
        return ''
    if name in SPECIAL_SHORT:
        return SPECIAL_SHORT[name]
    s = str(name)
    for suf in ['省', '市', '维吾尔自治区', '壮族自治区', '回族自治区', '自治区', '特别行政区']:
        s = s.replace(suf, '')
    return s[:3]

wb = load_workbook(source_file, data_only=True)
ws = wb['省份日历_人数底表']
source_meta = {}
for source_sheet in ['0811最新校历来源', '0805最新校历来源']:
    if source_sheet in wb.sheetnames:
        sws = wb[source_sheet]
        header = [sws.cell(1, c).value for c in range(1, sws.max_column + 1)]
        def _col(name):
            return header.index(name) + 1 if name in header else None
        pc, gc, sc = _col('省份'), _col('年级'), _col('信息来源')
        if pc and gc and sc:
            for rr in range(2, sws.max_row + 1):
                pv = sws.cell(rr, pc).value
                gv = sws.cell(rr, gc).value
                sv = sws.cell(rr, sc).value
                if pv and gv:
                    source_meta[(str(pv), str(gv))] = str(sv or '').strip()
        break

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v)[:10] if v else ''

records = []
for r in range(2, ws.max_row + 1):
    province = ws.cell(r, 2).value
    if not province:
        continue
    item = {'province': province, 'short': short_province(province), 'region': ws.cell(r, 1).value or '', 'grades': {}}
    c = 3
    for g in GRADE_ORDER:
        item['grades'][g] = {
            'holiday': fmt_date(ws.cell(r, c).value),
            'spring': fmt_date(ws.cell(r, c + 1).value),
            'isReal': ws.cell(r, c + 2).value or '否',
            'count': int(ws.cell(r, c + 3).value or 0),
            'pct': float(ws.cell(r, c + 4).value or 0),
            'source': source_meta.get((str(province), g), ''),
        }
        c += 5
    records.append(item)

def build_dist(event_key):
    buckets = defaultdict(lambda: {'count': 0, 'provinceSet': set()})
    for rec in records:
        for g in GRADE_ORDER:
            gd = rec['grades'][g]
            if gd[event_key]:
                buckets[gd[event_key]]['count'] += gd['count']
                buckets[gd[event_key]]['provinceSet'].add(rec['province'])
    total = sum(v['count'] for v in buckets.values())
    cum = 0
    rows = []
    for d in sorted(buckets):
        cnt = buckets[d]['count']
        cum += cnt
        rows.append({
            'date': d,
            'count': cnt,
            'share': cnt / total if total else 0,
            'cumulative': cum,
            'cumulativeShare': cum / total if total else 0,
            'provinceCount': len(buckets[d]['provinceSet']),
        })
    return {'rows': rows, 'total': total}

def build_grade_groups(event_key):
    out = {}
    for g in GRADE_ORDER:
        buckets = defaultdict(list)
        total = 0
        for rec in records:
            gd = rec['grades'][g]
            if gd[event_key]:
                buckets[gd[event_key]].append({'province': rec['province'], 'short': rec['short'], 'count': gd['count'], 'pct': gd['pct']})
                total += gd['count']
        rows = []
        for d in sorted(buckets):
            items = sorted(buckets[d], key=lambda x: (-x['count'], x['province']))
            cnt = sum(x['count'] for x in items)
            rows.append({'date': d, 'items': items, 'dateTotal': cnt, 'dateShare': cnt / total if total else 0})
        out[g] = {'rows': rows, 'total': total}
    return out

def build_top5():
    out = {}
    for g in GRADE_ORDER:
        total = sum(rec['grades'][g]['count'] for rec in records)
        arr = []
        for rec in records:
            gd = rec['grades'][g]
            arr.append({
                'grade': g,
                'province': rec['province'],
                'short': rec['short'],
                'count': gd['count'],
                'pct': gd['count'] / total if total else 0,
                'holiday': gd['holiday'],
                'spring': gd['spring'],
                'isReal': gd['isReal'],
            })
        out[g] = sorted(arr, key=lambda x: (-x['count'], x['province']))[:5]
    return out


def build_dist_for_grades(event_key, grades):
    buckets = defaultdict(lambda: {'count': 0, 'provinceSet': set()})
    for rec in records:
        for g in grades:
            gd = rec['grades'][g]
            if gd[event_key]:
                buckets[gd[event_key]]['count'] += gd['count']
                buckets[gd[event_key]]['provinceSet'].add(rec['province'])
    total = sum(v['count'] for v in buckets.values())
    cum = 0
    rows = []
    for d in sorted(buckets):
        cnt = buckets[d]['count']
        cum += cnt
        rows.append({'date': d, 'count': cnt, 'share': cnt / total if total else 0, 'cumulative': cum, 'cumulativeShare': cum / total if total else 0, 'provinceCount': len(buckets[d]['provinceSet'])})
    return {'rows': rows, 'total': total}

def build_grade_groups_for_grades(event_key, grades):
    all_groups = build_grade_groups(event_key)
    return {g: all_groups[g] for g in grades}

def build_top5_for_grades(grades):
    all_top5 = build_top5()
    return {g: all_top5[g] for g in grades}

payload = {
    'records': records,
    'stageGrades': STAGE_GRADES,
    'dist': {stage: {'holiday': build_dist_for_grades('holiday', grades), 'spring': build_dist_for_grades('spring', grades)} for stage, grades in STAGE_GRADES.items()},
    'gradeGroups': {stage: {'holiday': build_grade_groups_for_grades('holiday', grades), 'spring': build_grade_groups_for_grades('spring', grades)} for stage, grades in STAGE_GRADES.items()},
    'top5': {stage: build_top5_for_grades(grades) for stage, grades in STAGE_GRADES.items()},
    'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
}
json_data = json.dumps(payload, ensure_ascii=False)
geojson_path = Path(__file__).resolve().parents[1] / 'data' / 'china-mainland-trimmed.json'
geojson_data = geojson_path.read_text(encoding='utf-8')

html = r'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>有道升学-正价学员放假&开学时间分布</title>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.1/dist/echarts.min.js"></script>
<style>
:root{--bg:#f5f7fb;--card:#fff;--ink:#1f2630;--muted:#738096;--blue:#4f72f6;--blue2:#3e66e4;--bar:#75b9f3;--purple:#7b42ef;--line:#e6edf7;--shadow:0 16px 48px rgba(46,73,123,.10)}
*{box-sizing:border-box}body{margin:0;background:var(--bg);font-family:"PingFang SC","Microsoft YaHei",sans-serif;color:var(--ink)}
.shell{max-width:1840px;margin:0 auto;padding:22px 36px 50px}.header{position:relative;min-height:108px;display:flex;align-items:center;justify-content:center}.refresh-note{position:absolute;left:0;top:10px;color:#6f7f99;font-size:18px;font-weight:700}.title{font-size:58px;line-height:1;white-space:nowrap;color:#5975f6;font-weight:900;letter-spacing:-.04em;margin:0}.youdao-logo{position:absolute;right:0;top:0;width:540px;height:72px;display:flex;align-items:center;justify-content:flex-end}.youdao-logo img{width:540px;height:auto;display:block}.section-head{margin:26px 0 14px}.section-head h2{margin:0 0 16px;color:#5475f8;font-size:38px;font-weight:900}.section-head-line{display:flex;align-items:center;justify-content:space-between;gap:20px;margin-bottom:16px}.section-controls{display:flex;gap:10px;align-items:center}.rule{height:4px;background:#5a78f6;border-radius:99px}.toolbar-card,.kpis,.panel{border-radius:24px;background:rgba(255,255,255,.74);border:1px solid #dfe8fb;box-shadow:var(--shadow)}.toolbar-card{padding:18px 22px}.toolbar{display:grid;grid-template-columns:300px 260px 420px 1fr 260px;gap:18px;align-items:center}.ctrl{height:58px;background:#fff;border-radius:12px;border:1px solid #edf1f8;padding:0 18px;display:flex;align-items:center;justify-content:space-between;font-size:18px;width:300px}.ctrl b{font-size:19px}.badge{display:inline-flex;margin-left:8px;min-width:28px;height:28px;align-items:center;justify-content:center;border-radius:50%;background:#eef2ff;color:#4f72f6}.toolbar-note{font-size:16px;color:#738096;font-weight:700}.upload-btn{height:64px;background:#eef2f7;border:2px solid #cbd5e1;border-radius:14px;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#1d4ed8;font-size:19px;font-weight:900;cursor:pointer;box-shadow:0 10px 24px rgba(71,85,105,.14);line-height:1.15}.upload-btn small{margin-top:5px;color:#64748b;font-size:12px;font-weight:800}.upload-btn:hover{background:#e2e8f0;border-color:#94a3b8;transform:translateY(-1px)}.upload-status{font-size:13px;color:#738096;text-align:right}.sync-status{margin-top:14px;padding:12px 16px;border-radius:14px;border:1px solid #dfe8fb;background:#f8fbff;color:#64748b;font-size:15px;font-weight:800;display:flex;align-items:center;gap:10px}.sync-status:before{content:'';width:10px;height:10px;border-radius:50%;background:#94a3b8;display:inline-block}.sync-status.success{background:#f0fdf4;border-color:#bbf7d0;color:#15803d}.sync-status.success:before{background:#22c55e}.sync-status.error{background:#fef2f2;border-color:#fecaca;color:#b91c1c}.sync-status.error:before{background:#ef4444}.sync-status.loading{background:#eff6ff;border-color:#bfdbfe;color:#1d4ed8}.sync-status.loading:before{background:#3b82f6;box-shadow:0 0 0 4px rgba(59,130,246,.12)}.pill{display:flex;width:max-content;gap:6px;background:#eef3ff;border:1px solid #dfe8fb;border-radius:999px;padding:5px}.pill button{border:0;background:transparent;padding:10px 30px;border-radius:999px;font-size:18px;font-weight:900;color:#59677d;cursor:pointer}.pill button.active{background:#fff;color:#315be4;box-shadow:0 8px 20px rgba(79,114,246,.15)}.grade-pill{max-width:420px;overflow:auto}.grade-pill button{padding:10px 20px;white-space:nowrap}.kpis{padding:22px;margin-top:22px}.kpi-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:18px}.kpi{background:#fff;border-radius:16px;min-height:124px;display:flex;flex-direction:column;align-items:center;justify-content:center;border:1px solid #eff3fa}.kpi .label{font-size:22px;font-weight:900;margin-bottom:15px}.kpi .value{font-size:50px;color:#5b7cff;font-weight:300;line-height:1}.kpi .value.small{font-size:30px;font-weight:500}.main{margin-top:28px}.panel{padding:20px;background:#f7f8fb}.dist-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px}.card{background:#fff;border:1px solid #edf2f7;border-radius:14px;overflow:hidden;box-shadow:0 6px 20px rgba(73,86,112,.05)}.card-head{padding:18px 24px 14px;border-bottom:1px solid #edf2f7;font-size:25px;font-weight:900}.card-body{padding:14px 18px}.bar-chart{height:auto;padding:4px 2px}.bar-row{display:grid;grid-template-columns:104px 1fr;align-items:center;height:30px;gap:10px}.bar-date{font-size:14px;color:#666;text-align:left}.bar-track{height:22px;position:relative;background:transparent;display:flex;justify-content:center}.bar-fill{height:22px;background:#76b9f2;min-width:2px;display:flex;align-items:center;justify-content:center;color:#24506e;font-weight:500;font-size:14px}.bar-fill.purple{background:#874cf3;color:#fff}.table-wrap{max-height:380px;overflow:auto;border-radius:10px}table{width:100%;border-collapse:collapse;font-size:15px}th{position:sticky;top:0;background:#3e66e4;color:#fff;padding:11px 8px;z-index:1;border-right:1px solid rgba(255,255,255,.3);white-space:nowrap}td{padding:10px 8px;text-align:center;border-bottom:1px solid #eef2f7;color:#555;white-space:nowrap}tbody tr:nth-child(even) td{background:#f7f7f9}td.date,td.blue{background:#3e66e4!important;color:#fff!important;font-weight:800}.date-cell{width:96px!important;min-width:96px!important;max-width:96px!important;padding-left:6px!important;padding-right:6px!important}.province-cell{width:72px!important;min-width:72px!important;max-width:72px!important;padding-left:6px!important;padding-right:6px!important}.pct,.strong{font-weight:900;color:#555}.split{height:16px}.top5-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px}.top5-card .card-head{font-size:23px}.top5-rank{display:inline-flex;width:24px;height:24px;border-radius:50%;align-items:center;justify-content:center;background:#eef3ff;color:#315be4;font-weight:900}.map-layout{display:grid;grid-template-columns:1.35fr 1fr;gap:24px;align-items:stretch}.map-card{background:#fff;border:1px solid #edf2f7;border-radius:14px;padding:20px;box-shadow:0 6px 20px rgba(73,86,112,.05)}.map-card-head{display:flex;align-items:center;gap:8px;color:#34405a;font-size:16px;font-weight:900;margin-bottom:8px}.dot{width:8px;height:8px;border-radius:99px;background:#5b7cff;display:inline-block}.china-map{height:620px;width:100%;background:#fff}.rank-chart{height:620px;width:100%;background:#fff}.map-fallback{height:620px;display:flex;align-items:center;justify-content:center;color:#315be4;font-size:18px;font-weight:900}.map-note{text-align:center;color:#34405a;font-size:16px;font-weight:900;margin:0 0 8px}.map-note span{color:#315be4}.map-filter-row{display:flex;justify-content:flex-end;margin-bottom:16px}.stage-title{margin:34px 0 14px;display:flex;align-items:flex-end;justify-content:space-between}.stage-title h2{margin:0;color:#5475f8;font-size:36px;font-weight:900}.stage-title span{color:#8a98ad}.grade-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px}.grade-pair{display:flex;flex-direction:column;gap:20px}.grade-card .card-head{font-size:23px}.grade-note{padding:0 20px 8px;color:#8090a8;font-size:13px}@media(max-width:1300px){.title{font-size:42px}.row3,.top5-grid,.kpi-grid{grid-template-columns:repeat(2,1fr)}.map-layout,.dist-grid{grid-template-columns:1fr}.china-map,.rank-chart{height:560px}}@media(max-width:820px){.title{font-size:30px;white-space:normal}.header{justify-content:flex-start;padding-top:72px}.youdao-logo{left:0;right:auto}.row3,.top5-grid,.grade-grid,.grade-pair,.kpi-grid,.dist-grid{grid-template-columns:1fr}.toolbar{grid-template-columns:1fr}.section-head-line{display:block}.section-controls{margin-top:12px;overflow:auto}.china-map,.rank-chart{height:460px}}
</style>
</head>
<body>
<div class="shell">
  <header class="header">
    <div class="refresh-note">每天10点自动刷新官方校历</div>
    <h1 class="title">有道升学-正价学员放假&开学时间分布</h1>
  </header>

  <div class="toolbar-card"><div class="toolbar">
    <div class="ctrl">年份:<span><b>2027</b><span class="badge">1</span></span></div>
    <div class="toolbar-note" id="uploadStatus">当前口径：正价在班学员，按省份与年级去重后汇总</div>
    <div></div><div></div>
    <label class="upload-btn"><span>上传在班人数</span><small>支持 CSV / XLSX / XLS</small><input id="studentUpload" type="file" accept=".csv,.xlsx,.xls" style="display:none"></label>
  </div><div id="syncStatus" class="sync-status">当前状态：默认数据已加载。通过本地服务上传后，会写回本地文件并同步 GitHub Pages。</div></div>

  <main class="main">
    <section class="section-head"><div class="section-head-line"><h2>①放假&开学时间分布图</h2><div class="section-controls"><div class="pill" id="stageTabsOne"><button data-stage="初中">初中</button><button data-stage="高中" class="active">高中</button></div><div class="pill grade-pill" id="gradeTabsOne"></div></div></div><div class="rule"></div></section>
    <div class="kpis"><div class="kpi-grid">
      <div class="kpi"><div class="label">覆盖省份数</div><div class="value" id="kpiProvince">31</div></div>
      <div class="kpi"><div class="label">官方校历发布省份数</div><div class="value" id="kpiOfficial">0</div></div>
      <div class="kpi"><div class="label">最早放假时间</div><div class="value small" id="kpiHMin">-</div></div>
      <div class="kpi"><div class="label">最晚放假时间</div><div class="value small" id="kpiHMax">-</div></div>
      <div class="kpi"><div class="label">最早开学时间</div><div class="value small" id="kpiSMin">-</div></div>
      <div class="kpi"><div class="label">最晚开学时间</div><div class="value small" id="kpiSMax">-</div></div>
    </div></div>
    <div class="panel">
      <div class="dist-grid">
        <div class="card"><div class="card-head">寒假放假时间聚合分布</div><div class="card-body"><div id="holidayBars" class="bar-chart"></div></div></div>
        <div class="card"><div class="card-head">春季开学时间聚合分布</div><div class="card-body"><div id="springBars" class="bar-chart"></div></div></div>
      </div>
    </div>

    <section class="section-head"><div class="section-head-line"><h2>②在班学员所在省份分布图</h2><div class="section-controls"><div class="pill" id="stageTabsTwo"><button data-stage="初中">初中</button><button data-stage="高中" class="active">高中</button></div><div class="pill grade-pill" id="gradeTabsTwo"></div></div></div><div class="rule"></div></section>
    <div class="panel">
      <div class="map-layout">
        <div class="map-card"><div class="map-card-head"><span class="dot"></span>省份分布地图（颜色越深人数越多）</div><div id="chinaMap" class="china-map"></div></div>
        <div class="map-card"><div class="map-card-head"><span class="dot"></span>省份TOP排名（柱状图）</div><div class="map-note"><span id="mapGradeLabel">高一</span>－在班学员总数: <span id="mapTotal">0</span> 人</div><div id="provinceRankChart" class="rank-chart"></div></div>
      </div>
    </div>

    <section class="section-head"><div class="section-head-line"><h2>③各年级分省份时间分布情况</h2><span class="toolbar-note">按当前学段展示各年级寒假放假 / 春季开学明细</span></div><div class="rule"></div></section>
    <div class="grade-grid" id="gradeGrid"></div>
  </main>
</div>
<script>
const DATA = __DATA__;
const CHINA_MAINLAND_GEO = __GEOJSON__;
let stage = '高中';
let gradeFilter = '高一';
let DERIVED = null;
const fmtNum = n => (n || 0).toLocaleString('zh-CN');
const fmtPct = n => ((n || 0) * 100).toFixed(2) + '%';
const grades = () => DATA.stageGrades[stage];
const selectedGrades = () => gradeFilter === '全部' ? grades() : [gradeFilter];
function setSyncStatus(type, text){const el=document.getElementById('syncStatus'); if(!el)return; el.className='sync-status '+(type||''); el.textContent=text;}
const suffixes = ['省','市','维吾尔自治区','壮族自治区','回族自治区','自治区','特别行政区'];
function normProvince(v){let s=String(v||'').trim(); suffixes.forEach(x=>s=s.replaceAll(x,'')); return s;}
function normGrade(v){let s=String(v||'').trim(); const m={'七年级':'初一','八年级':'初二','九年级':'初三','高一年级':'高一','高二年级':'高二','高三年级':'高三'}; return m[s]||s;}
function recompute(){
  const dist={}, gradeGroups={}, top5={};
  Object.entries(DATA.stageGrades).forEach(([st,gs])=>{
    dist[st]={}; gradeGroups[st]={holiday:{},spring:{}}; top5[st]={};
    ['holiday','spring'].forEach(event=>{
      const buckets={}; gs.forEach(g=>DATA.records.forEach(rec=>{const gd=rec.grades[g]; const d=gd[event]; if(!d)return; if(!buckets[d])buckets[d]={count:0,provinceSet:new Set()}; buckets[d].count+=gd.count; buckets[d].provinceSet.add(rec.province)}));
      const total=Object.values(buckets).reduce((a,b)=>a+b.count,0); let cum=0;
      dist[st][event]={total,rows:Object.keys(buckets).sort().map(d=>{const cnt=buckets[d].count; cum+=cnt; return {date:d,count:cnt,share:total?cnt/total:0,cumulative:cum,cumulativeShare:total?cum/total:0,provinceCount:buckets[d].provinceSet.size}})};
      gs.forEach(g=>{const gb={}; let gtotal=0; DATA.records.forEach(rec=>{const gd=rec.grades[g]; const d=gd[event]; if(!d)return; if(!gb[d])gb[d]=[]; gb[d].push({province:rec.province,short:rec.short,count:gd.count,pct:gd.pct}); gtotal+=gd.count}); gradeGroups[st][event][g]={total:gtotal,rows:Object.keys(gb).sort().map(d=>{const items=gb[d].sort((a,b)=>b.count-a.count||a.province.localeCompare(b.province,'zh-CN')); const dateTotal=items.reduce((a,b)=>a+b.count,0); return {date:d,items,dateTotal,dateShare:gtotal?dateTotal/gtotal:0}})};});
    });
    gs.forEach(g=>{const total=DATA.records.reduce((a,r)=>a+r.grades[g].count,0); top5[st][g]=DATA.records.map(r=>({province:r.province,short:r.short,count:r.grades[g].count,pct:total?r.grades[g].count/total:0,holiday:r.grades[g].holiday,spring:r.grades[g].spring})).sort((a,b)=>b.count-a.count||a.province.localeCompare(b.province,'zh-CN')).slice(0,5)});
  });
  DERIVED={dist,gradeGroups,top5};
}
function buildEventDist(event, gs){const buckets={}; gs.forEach(g=>DATA.records.forEach(rec=>{const gd=rec.grades[g]; const d=gd[event]; if(!d)return; if(!buckets[d])buckets[d]={count:0,provinceSet:new Set()}; buckets[d].count+=gd.count; buckets[d].provinceSet.add(rec.province)})); const total=Object.values(buckets).reduce((a,b)=>a+b.count,0); let cum=0; return {total,rows:Object.keys(buckets).sort().map(d=>{const cnt=buckets[d].count; cum+=cnt; return {date:d,count:cnt,share:total?cnt/total:0,cumulative:cum,cumulativeShare:total?cum/total:0,provinceCount:buckets[d].provinceSet.size}})}}
function viewDist(event){return gradeFilter === '全部' ? DERIVED.dist[stage][event] : buildEventDist(event, selectedGrades())}
function dateRange(event){const arr=[];DATA.records.forEach(r=>selectedGrades().forEach(g=>{const d=r.grades[g][event];if(d)arr.push(d)}));arr.sort();return [arr[0]||'-',arr[arr.length-1]||'-']}
function updateKpi(){const ps=new Set(), official=new Set();DATA.records.forEach(r=>selectedGrades().forEach(g=>{ps.add(r.province); const src=(r.grades[g].source||'').trim(); if(src && !src.includes('待官方发布'))official.add(r.province)}));const hr=dateRange('holiday'),sr=dateRange('spring');document.getElementById('kpiProvince').textContent=ps.size;document.getElementById('kpiOfficial').textContent=official.size;document.getElementById('kpiHMin').textContent=hr[0];document.getElementById('kpiHMax').textContent=hr[1];document.getElementById('kpiSMin').textContent=sr[0];document.getElementById('kpiSMax').textContent=sr[1]}
function renderTables(event,distId,aggId,label){const rows=viewDist(event).rows;document.getElementById(distId).innerHTML='<thead><tr><th>'+label+'</th><th>在班人数</th><th>在班占比</th></tr></thead><tbody>'+rows.map(r=>`<tr><td class="date date-cell">${r.date}</td><td>${fmtNum(r.count)}</td><td class="pct">${fmtPct(r.share)}</td></tr>`).join('')+'</tbody>';document.getElementById(aggId).innerHTML='<thead><tr><th>'+label+'</th><th>累计人数</th><th>累计百分比</th></tr></thead><tbody>'+rows.map(r=>`<tr><td class="date date-cell">${r.date}</td><td>${fmtNum(r.cumulative)}</td><td class="pct">${fmtPct(r.cumulativeShare)}</td></tr>`).join('')+'</tbody>'}
function renderBars(id,event,colorClass=''){const rows=viewDist(event).rows;document.getElementById(id).innerHTML=rows.map(r=>`<div class="bar-row"><div class="bar-date">${r.date}</div><div class="bar-track"><div class="bar-fill ${colorClass}" style="width:${Math.max(0.8,r.cumulativeShare*100)}%">${fmtPct(r.cumulativeShare)}</div></div></div>`).join('')}
function renderTop5(){document.getElementById('top5Grid').innerHTML=grades().map(g=>`<div class="card top5-card"><div class="card-head">${g}在班省份Top5</div><div class="card-body"><table><thead><tr><th>排名</th><th class="province-cell">省份</th><th>在班人数</th><th>占比</th><th>放假时间</th><th>开学时间</th></tr></thead><tbody>${DERIVED.top5[stage][g].map((x,i)=>`<tr><td><span class="top5-rank">${i+1}</span></td><td class="blue province-cell">${x.short}</td><td class="strong">${fmtNum(x.count)}</td><td class="pct">${fmtPct(x.pct)}</td><td>${x.holiday}</td><td>${x.spring}</td></tr>`).join('')}</tbody></table></div></div>`).join('')}
function gradeTable(g,event,label){const groups=DERIVED.gradeGroups[stage][event][g].rows;let html=`<div class="table-wrap"><table><thead><tr><th>${label}</th><th class="province-cell">省份</th><th>在班人数</th><th>在班占比</th></tr></thead><tbody>`;groups.forEach(gr=>{const span=gr.items.length+1;gr.items.forEach((it,idx)=>{html+='<tr>';if(idx===0)html+=`<td class="date date-cell" rowspan="${span}">${gr.date}</td>`;html+=`<td class="blue province-cell">${it.short}</td><td class="strong">${fmtNum(it.count)}</td><td class="pct">${fmtPct(it.pct)}</td></tr>`});html+=`<tr><td class="blue province-cell">总计</td><td class="strong">${fmtNum(gr.dateTotal)}</td><td class="pct">${fmtPct(gr.dateShare)}</td></tr>`});return html+'</tbody></table></div>'}
function renderGrades(){document.getElementById('gradeGrid').innerHTML=grades().map(g=>`<div class="grade-pair"><div class="card grade-card"><div class="card-head">${g}寒假放假时间分布</div><div class="grade-note">省份简称展示，列宽已压缩</div><div class="card-body">${gradeTable(g,'holiday','寒假放假时间')}</div></div><div class="card grade-card"><div class="card-head">${g}春季开学时间分布</div><div class="grade-note">省份简称展示，列宽已压缩</div><div class="card-body">${gradeTable(g,'spring','春季开学时间')}</div></div></div>`).join('')}

function renderGradeTabs(){['gradeTabsOne','gradeTabsTwo'].forEach(id=>{const tabs=document.getElementById(id); if(!tabs)return; tabs.innerHTML=grades().map(g=>`<button data-grade="${g}" class="${gradeFilter===g?'active':''}">${g}</button>`).join(''); tabs.querySelectorAll('button').forEach(b=>b.onclick=()=>{gradeFilter=b.dataset.grade; render();});}); document.querySelectorAll('#stageTabsOne button,#stageTabsTwo button').forEach(b=>b.classList.toggle('active', b.dataset.stage===stage));}
function render(){recompute();renderGradeTabs();updateKpi();renderBars('holidayBars','holiday');renderBars('springBars','spring','purple');renderMap();renderGrades();}

const NAME_ALIAS={'北京市':'北京','天津市':'天津','河北省':'河北','山西省':'山西','内蒙古自治区':'内蒙古','辽宁省':'辽宁','吉林省':'吉林','黑龙江省':'黑龙江','上海市':'上海','江苏省':'江苏','浙江省':'浙江','安徽省':'安徽','福建省':'福建','江西省':'江西','山东省':'山东','河南省':'河南','湖北省':'湖北','湖南省':'湖南','广东省':'广东','广西壮族自治区':'广西','海南省':'海南','重庆市':'重庆','四川省':'四川','贵州省':'贵州','云南省':'云南','西藏自治区':'西藏','陕西省':'陕西','甘肃省':'甘肃','青海省':'青海','宁夏回族自治区':'宁夏','新疆维吾尔自治区':'新疆'};
let chinaMapReady=null,mapChart=null,rankChart=null;
function ensureChinaMap(){if(!window.echarts)return Promise.resolve(false); if(chinaMapReady)return chinaMapReady; chinaMapReady=Promise.resolve().then(()=>{echarts.registerMap('china-mainland',CHINA_MAINLAND_GEO); return true;}).catch(err=>{console.error('mainland map register failed',err); return false;}); return chinaMapReady;}
function provinceRows(){const total=DATA.records.reduce((a,r)=>a+(r.grades[gradeFilter]?.count||0),0);return DATA.records.map(r=>({name:r.province,fullName:r.province,short:r.short,value:r.grades[gradeFilter].count,count:r.grades[gradeFilter].count,pct:total?r.grades[gradeFilter].count/total:0})).sort((a,b)=>b.count-a.count||a.fullName.localeCompare(b.fullName,'zh-CN'))}
async function renderMap(){const rows=provinceRows();const total=rows.reduce((a,b)=>a+b.count,0);document.getElementById('mapGradeLabel').textContent=gradeFilter;document.getElementById('mapTotal').textContent=fmtNum(total);const max=Math.max(...rows.map(x=>x.count),1);const rankEl=document.getElementById('provinceRankChart');rankChart=echarts.getInstanceByDom(rankEl)||echarts.init(rankEl);const rankRows=rows.filter(x=>x.count>0);rankChart.setOption({backgroundColor:'#fff',grid:{left:70,right:50,top:20,bottom:28},tooltip:{trigger:'axis',axisPointer:{type:'shadow'}},xAxis:{type:'value',axisLine:{show:false},axisTick:{show:false},splitLine:{lineStyle:{color:'#eef2f7'}},axisLabel:{color:'#7a8494'}},yAxis:{type:'category',inverse:true,data:rankRows.map(x=>x.short),axisLine:{show:false},axisTick:{show:false},axisLabel:{color:'#4b5563',fontWeight:800}},series:[{type:'bar',data:rankRows.map((x,i)=>({value:x.count,itemStyle:{color:i<3?'#4d3fd8':'#b9c8ff'}})),barWidth:12,label:{show:true,position:'right',color:'#5b6472',fontSize:10}}]},true);const ok=await ensureChinaMap();const mapEl=document.getElementById('chinaMap');if(!ok){mapEl.innerHTML='<div class="map-fallback">地图注册失败，但右侧省份排名已正常显示</div>';setTimeout(()=>{rankChart&&rankChart.resize();},0);return;}mapEl.innerHTML='';mapChart=echarts.getInstanceByDom(mapEl)||echarts.init(mapEl);mapChart.setOption({backgroundColor:'#fff',tooltip:{trigger:'item',backgroundColor:'#fff',borderColor:'#dbeafe',borderWidth:1,extraCssText:'box-shadow:0 10px 28px rgba(49,91,228,.18);border-radius:12px;',textStyle:{color:'#24437a',fontSize:14,fontWeight:800},formatter:p=>{const d=p.data||{};return `${d.short||p.name}<br/>在班人数 <b style="color:#315be4">${fmtNum(d.count||0)}</b><br/>在班占比 <b style="color:#315be4">${fmtPct(d.pct||0)}</b>`}},visualMap:{show:true,min:0,max:max,left:28,bottom:46,text:['多','少'],itemHeight:150,itemWidth:16,inRange:{color:['#e8efff','#9fb5ff','#4d3fd8']},textStyle:{color:'#64748b',fontWeight:800}},series:[{type:'map',map:'china-mainland',roam:false,zoom:1.08,top:16,bottom:16,left:0,right:0,label:{show:true,formatter:p=>(p.data&&p.data.short)||'',color:'#536070',fontSize:9,fontWeight:700},emphasis:{label:{show:true,formatter:p=>(p.data&&p.data.short)||p.name,color:'#1d4ed8',fontWeight:900,fontSize:10},itemStyle:{areaColor:'#eef2ff',borderColor:'#4d3fd8',borderWidth:1.6}},itemStyle:{borderColor:'#e5e7eb',borderWidth:1,areaColor:'#eef2ff'},data:rows}]},true);mapChart.off('click');mapChart.on('click',p=>{mapChart.dispatchAction({type:'showTip',seriesIndex:0,name:p.name});});setTimeout(()=>{mapChart&&mapChart.resize();rankChart&&rankChart.resize();},0)}
window.addEventListener('resize',()=>{mapChart&&mapChart.resize();rankChart&&rankChart.resize();});

function applyUploadedRows(rows, filename){
  const status=document.getElementById('uploadStatus');
  const provinceMap={}; DATA.records.forEach(r=>{provinceMap[normProvince(r.province)]=r; provinceMap[normProvince(r.short)]=r});
  const seen=new Set(); const counts={}; let valid=0, skipped=0;
  rows.forEach(row=>{const lower={}; Object.keys(row).forEach(k=>lower[String(k).trim().toLowerCase()]=row[k]); const uid=lower.userid||lower.user_id||lower.uid||lower['用户id']||lower['学员id']; const grade=normGrade(lower.gradename||lower.grade||lower['年级']); const rawProvince=lower.province||lower['省份']||lower['省']; const rec=provinceMap[normProvince(rawProvince)]; if(!uid||!grade||!rec||!DATA.stageGrades['初中'].concat(DATA.stageGrades['高中']).includes(grade)){skipped++; return;} const key=`${uid}|${grade}|${rec.province}`; if(seen.has(key))return; seen.add(key); counts[`${grade}|${rec.province}`]=(counts[`${grade}|${rec.province}`]||0)+1; valid++;});
  const gradeTotals={}; DATA.stageGrades['初中'].concat(DATA.stageGrades['高中']).forEach(g=>gradeTotals[g]=0);
  DATA.records.forEach(r=>Object.keys(gradeTotals).forEach(g=>{const c=counts[`${g}|${r.province}`]||0; r.grades[g].count=c; gradeTotals[g]+=c;}));
  DATA.records.forEach(r=>Object.keys(gradeTotals).forEach(g=>{r.grades[g].pct=gradeTotals[g]?r.grades[g].count/gradeTotals[g]:0;}));
  status.textContent=`已上传：${filename}｜有效去重 ${fmtNum(valid)} 条｜跳过 ${fmtNum(skipped)} 条`;
  render();
}
document.getElementById('studentUpload').addEventListener('change', async e=>{
  const file=e.target.files[0]; if(!file)return;
  const status=document.getElementById('uploadStatus');
  const isLocalService = ['127.0.0.1','localhost'].includes(location.hostname) && location.protocol.startsWith('http');
  if(isLocalService){
    try{
      status.textContent='正在上传并写回本地文件...';
      setSyncStatus('loading','正在处理：解析上传文件 → 写回本地 Excel → 重新生成看板 → 同步 GitHub Pages...');
      const form=new FormData(); form.append('file',file);
      const resp=await fetch('/api/upload-students',{method:'POST',body:form});
      const result=await resp.json().catch(()=>({error:'服务返回异常，无法解析结果'}));
      if(resp.ok && result.ok){
        const pub=result.publish||{};
        if(pub.git==='pushed'){
          status.textContent=`已写回并同步：${result.filename}｜有效去重 ${fmtNum(result.valid)} 条｜跳过 ${fmtNum(result.skipped)} 条`;
          setSyncStatus('success',`同步成功：本地 Excel 与 GitHub Pages 已更新。线上链接通常几十秒内刷新：${pub.url||'https://yxq870504-lgtm.github.io/SXZX-FJKXSJ/'}`);
        }else if(pub.git==='unchanged'){
          status.textContent=`已写回本地文件：${result.filename}｜GitHub 内容无变化`;
          setSyncStatus('success','处理成功：本地文件已写回，index.html 无变化，本次无需推送 GitHub。');
        }else{
          status.textContent=`已写回本地文件：${result.filename}｜GitHub 未同步`;
          setSyncStatus('error',`本地写回成功，但 GitHub 未同步：${pub.message||'未知原因'}`);
        }
        setTimeout(()=>location.reload(),1200);
        e.target.value='';
        return;
      }
      status.textContent='上传处理失败';
      setSyncStatus('error',`同步失败：${result.error||result.message||'请查看本地服务窗口日志'}`);
      e.target.value='';
      return;
    }catch(err){
      status.textContent='上传处理失败';
      setSyncStatus('error',`同步失败：${err.message||err}。请确认 start_dashboard_server.bat 服务窗口仍在运行。`);
      e.target.value='';
      return;
    }
  }
  const buf=await file.arrayBuffer();
  const wb=XLSX.read(buf,{type:'array'});
  const ws=wb.Sheets[wb.SheetNames[0]];
  const rows=XLSX.utils.sheet_to_json(ws,{defval:''});
  applyUploadedRows(rows,file.name);
  setSyncStatus('success','已在当前浏览器临时刷新。注意：线上 GitHub Pages 静态页面不能直接写回本地或推送 GitHub。');
  e.target.value='';
});

document.querySelectorAll('#stageTabsOne button,#stageTabsTwo button').forEach(b=>b.onclick=()=>{stage=b.dataset.stage;gradeFilter=grades()[0];render();});
render();
</script>
</body>
</html>'''.replace('__DATA__', json_data).replace('__GEOJSON__', geojson_data)

out_file.write_text(html, encoding='utf-8')
print(out_file)
